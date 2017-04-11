# -*- coding: utf-8 -*-
import json
import uuid

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, HttpResponseRedirect, StreamingHttpResponse, Http404
import numpy as np
import pandas as pd

from sklearn import preprocessing, linear_model
from openpyxl import Workbook


from . import models

from .zhuanlu import PRO_BOF_HIS_ALLFIELDS
from QinggangManageSys.settings import MAIN_OUTFIT_BASE,MEDIA_ROOT,MEDIA_URL

# def load(request):
#     if not request.user.is_authenticated():
#         return HttpResponseRedirect("/login")
#     contentVO={
#         'title':'数据分析工具',
#         'state':'success'
#     }
#     return render(request, 'data_import/analysis_tool.htm',contentVO)

def relation_ana(request):
    if not request.user.is_authenticated():
            return HttpResponseRedirect("/login")
    contentVO={
        'title':'数据分析工具——关联性分析',
        'state':'success'
    }

    return HttpResponse(json.dumps(contentVO), content_type='application/json')
def data_cleaning(data):
    return data.fillna(0)

def linear_regression_model(data):
    cols = list(data.columns)
    length_cols = len(cols)
    features = cols[:-1]
    out = cols[-1]

    data_array = np.array(data)
    scale_data_array = preprocessing.scale(data_array)
    data_scale_df = pd.DataFrame(scale_data_array,columns=cols)

    X_scale = data_scale_df[features]
    Y_scale = data_scale_df[out]

    regr = linear_model.LinearRegression()
    regr.fit(X_scale, Y_scale)

    return regr.coef_, regr.intercept_


def wushu_ana(df):
    """
    五数区间
    """
    Q1 = np.percentile(df,25)
    Q3 = np.percentile(df,75)
    L = 2*Q1 - Q3
    H = 2*Q3 - Q1
    ana_result={}
    ana_result['Q1'] = Q1
    ana_result['Q3'] = Q3
    ana_result['down'] = L
    ana_result['top'] = H
    return ana_result

def regression(output,selected_eles,db_table_name):
    """
    @param output 回归应变量 str
    @param selected_eles 回归自变量list
    @param db_table_name 存储回归结果的表
    @rtn coef 与selected_eles对应的回归系数list
    @rtn intercep 回归方程截距
    @warning: 如果清洗后无数据，返回false，注意处理这种情况
    """
    fout_des = open("data_number.txt", 'a+', encoding='utf-8')
    sqlVO = {"db_name": 'l2own'}

    isFiveAnalyse = dict()
    bound_lows = dict()
    bound_highs = dict()

    sql = 'select DATA_ITEM_EN,IF_FIVENUMBERSUMMARY,NUMERICAL_LOWER_BOUND,NUMERICAL_UPPER_BOUND from QG_USER.PRO_BOF_HIS_ALLSTRUCTURE WHERE  IF_ANALYSE_TEMP = 1'
    sqlVO["sql"] = sql
    rs = models.BaseManage().direct_select_query_orignal_sqlVO(sqlVO)

    for row in rs:

        isFiveAnalyse['%s' % row[0]] = '%s' % row[1]
        bound_lows['%s' % row[0]] = '%s' % row[2]
        bound_highs['%s' % row[0]] = '%s' % row[3]

    allcolumns = selected_eles + [output]
    columns_coma = ",".join(allcolumns)
    sql = 'SELECT ' + ', '.join(selected_eles) + ', ' + output + ' from QG_USER.PRO_BOF_HIS_ALLFIELDS'
    sqlVO["sql"] = sql
    rs = models.BaseManage().direct_select_query_orignal_sqlVO(sqlVO)
    """
    移入平台时需要将tuple类型的result转化为list
    """
    alldf = pd.DataFrame(list(rs), columns=allcolumns)
    # data_ready = data_cleaning(pd.DataFrame(rs, columns=allcolumns))
    five_downs = dict()
    five_highs = dict()

    for col in allcolumns:
        temp_df = alldf.copy()
        temp_df[col] = temp_df[col].dropna().map(lambda x:float(x))

        # """
        # 可添加 用均值对原数据集空值的填充,若不需要则注释
        # """
        # mean=temp_df[col].describe().get('mean',0)
        # # print('mean:',mean)
        # alldf[col] = alldf[col].fillna(mean)
        # """
        # end fill nan with mean
        # """

        #filter data by bound of low and high
        bound_low = float(bound_lows.get(col,-999999999999))
        bound_high = float(bound_highs.get(col,999999999999))
        temp_df = temp_df[(temp_df[col] >= bound_low) & ( temp_df[col] <= bound_high )]
        if isFiveAnalyse.get(col,'0') == '1':
            LH = wushu_ana(temp_df.sort_values(by=col)[col])
            five_downs['%s' % col ] = LH['down']
            five_highs['%s' % col ] = LH['top']
            print(col, five_downs[col], five_highs[col])


    alldf_temp = alldf.copy()
    # fout_des.write("%s before drop : %s\n" % (columns_coma,str(len(alldf))))
    before_drop_num = len(alldf)
    alldf = alldf.dropna(how='any')
    after_drop_num = len(alldf)
    fout_des.write("%s  nan rate : %s\n" % (columns_coma,str(before_drop_num/after_drop_num)))
    for col in allcolumns:
        temp_col_s = alldf[col]
        temp_col_s = temp_col_s.dropna()
        after_drop_temp_col = len(temp_col_s)
        fout_des.write("%s  nan rate : %s\n" % (col,str(after_drop_temp_col/after_drop_num)))
    """
    # 有字段的类型为Object
    for col in allcolumns:
        alldf[col] = alldf[col].map(lambda x:float(x))

    # 根据各因素上限联合筛选数据
    value_bound_tag = True
    for col in allcolumns:
        bound_low = float(bound_lows.get(col,-999999999999))
        bound_high = float(bound_highs.get(col,999999999999))
        alldf = alldf[(alldf[col] >= bound_low) & (alldf[col] <= bound_high)]
        if len(alldf) == 0:
            print("no data after bound")
            value_bound_tag = False
            break

    if  not value_bound_tag:
        return False
    # 五值分析
    five_tag = True
    for col in allcolumns:
        if isFiveAnalyse.get(col,'0') == '1':
            five_down = five_downs.get(col,-999999999999)
            five_high = five_highs.get(col,999999999999)
            alldf = alldf[( alldf[col] >= five_down ) & ( alldf[col] <= five_high )]
            if len(alldf) == 0:
                five_tag = False
                print("no data after five")
                break
    if not five_tag:
        return False

    coef, intercept = linear_regression_model(alldf)
    coef = list(map(lambda x: str(x), coef))
    """
    # save result to database
    """
    for i in range(len(selected_eles)):
        sql = 'insert into QG_USER.%s values(\'%s\', \'%s\', \'%s\')'%(db_table_name, output, selected_eles[i], coef[i])
        sqlVO['sql'] = sql
        models.BaseManage().direct_execute_query_sqlVO(sqlVO)
    sql = 'insert into QG_USER.%s values(\'%s\', \'BIAS\', \'%s\')'%(db_table_name, output, str(intercept))
    sqlVO['sql'] = sql
    models.BaseManage().direct_execute_query_sqlVO(sqlVO)
    """
    fout_des.close()
    # return coef, intercept



def regression_ana(result):
    pass


def report(request):
    if not request.user.is_authenticated():
            return HttpResponseRedirect("/login")
    contentVO={
        'title':'数据分析工具——生成分析结果报表',
        'state':'success'
    }
    db_table_names = {
        '321':"relation_cof_output_middle",
        '311':"relation_cof_output_input",
        '211':"relation_cof_middle_input",
        '322':"regression_cof_output_middle",
        '312':"regression_cof_output_input",
        '212':"regression_cof_middle_input",
    }
    sheetnames = {
        "relation_cof_output_middle":"相关系数_输出-控制",
        "relation_cof_output_input":"相关系数_输出-输入",
        "relation_cof_middle_input":"相关系数_控制-输入",
        "regression_cof_output_middle":"回归系数_输出-控制",
        "regression_cof_output_input":"回归系数_输出-输入",
        "regression_cof_middle_input":"回归系数_控制-输入",
    }
    entrance = [(3,2,1) ,(3,1,1), (2,1,1), (3,2,2) ,(3,1,2), (2,1,2)]


    filename = '相关性_回归分析汇总%s.xlsx' % str(uuid.uuid1())
    save_filename = MEDIA_ROOT + '/' + filename
    print(save_filename)
    contentVO['filepath'] = MEDIA_URL + filename

    sqlVO = dict()
    sqlVO["db_name"] = "l2own"

    wb = Workbook()
    i = 0

    for tags in entrance:
        tablekey = '%s%s%s' % tags
        table_name = db_table_names[tablekey]

        sql = 'SELECT * FROM QG_USER.%s' % table_name.upper()
        sqlVO["sql"] = sql

        rs = models.BaseManage().direct_select_query_orignal_sqlVO(sqlVO)
        if len(rs) == 0:
            continue

        if i == 0:
            ws = wb.active
            ws.title = sheetnames[table_name]
            i += 1
        else:
            ws = wb.create_sheet(sheetnames[table_name])

        rs_df = pd.DataFrame(list(rs))

        #sorted by column 1 and 3
        cols = list(rs_df.columns)
        rs_df['abs'] = rs_df[cols[2]].map(lambda x: abs(float(x)))
        # fout_sort = open('sort.txt', 'w+',encoding='utf-8')
        rs_df = rs_df.sort_values(by=[cols[0],'abs'],ascending=False)
        for ind in rs_df.index:
            temp = tuple(rs_df.loc[ind,cols])
            keyA, keyB, value = temp
            keyB = keyB.strip()
            if keyB == 'BIAS':
                nameB = "偏离-截距"
            else:
                nameB = PRO_BOF_HIS_ALLFIELDS[keyB]
            ws.append((PRO_BOF_HIS_ALLFIELDS[keyA.strip()], nameB, value))
            # fout_sort.write('%s,%s,%s\n'% tuple(rs_df.loc[ind,cols]))
        # fout_sort.close()
    print(save_filename)
    wb.save(save_filename)
    return HttpResponse(json.dumps(contentVO), content_type='application/json')
