#coding=UTF-8
import openpyxl
from QinggangManageSys.settings import BASE_DIR
from QinggangManageSys.settings import ROW_NUM

def get_primary_key(model):
    primary_key=None
    attrs_fields= model._meta.fields
    for field in attrs_fields:
        if field.primary_key:
            primary_key=field.attname
    return primary_key

def get_model_attrs(model):
	attrs_str= model.__doc__
	attrs_str=attrs_str[attrs_str.index('(')+1:attrs_str.index(')')]
	attrs_list=attrs_str.split(',')
	attrs_list_final=[]
	for each in attrs_list:
		if each.endswith("_ptr") or each=='id':
			continue
		each=each.strip(' ')
		attrs_list_final.append(each)
	print(attrs_list_final)
	return attrs_list_final



		
#通过传值的dictVO拼接成条件传入即‘SELECT * FROM WHERE ** and **’
def create_select_condition(dictVO):
	result={}
	varlist=[]
	whereCondition=" where "
	print(len(dictVO))
	if len(dictVO)>0:
		for each in dictVO:
			whereCondition+=each+'=%s'+' and '
			varlist.append(dictVO[each])
		whereCondition=whereCondition[:len(whereCondition)-1]+')'
		print('{0}'.format(whereCondition))
		result['whereCondition']=whereCondition
		result['vars']=varlist
	else:
		whereCondition+=" 1=1 "
		result['whereCondition']=whereCondition
	return result

def create_select_sqlVO(model,dictVO):
	condition = create_select_condition(dictVO)
	whereCondition=condition.get('whereCondition')
	if(model!=None):
		table_name = model._meta.db_table
	sql ='SELECT * FROM '+table_name+str(whereCondition)
	print(sql)
	return {'sql':sql,'vars':condition.get('vars')}



def create_insert_condition(dictVO):
	result={}
	varlist=[]
	columnCondition=" ("
	valueCondition=" ("
	for each in dictVO:
		columnCondition+=each+','
		ele = dictVO[each]
		if isinstance(ele,int) or isinstance(ele,float):
			valueCondition+='%s,'
		valueCondition+="'%s',"
		varlist.append(dictVO[each])
	columnCondition=columnCondition[:len(columnCondition)-1]+')'
	valueCondition=valueCondition[:len(valueCondition)-1]+')'
	result['columnCondition']=columnCondition
	result['valueCondition']=valueCondition
	result['vars']=varlist
	return result

def create_insert_sqlVO(model,dictVO):
	condition = create_insert_condition(dictVO)
	columnCondition=condition.get('columnCondition')
	valueCondition=condition.get('valueCondition')
	if(model!=None):
		table_name = model._meta.db_table
	sql ='INSERT INTO '+table_name+columnCondition+' VALUES'+valueCondition
	return {'sql':sql,'vars':condition.get('vars')}

#UODATE语句只能通过own_uid（即关联字段,同时也是每个表的主键字段）来更新
def create_update_condition(dictVO,primary_key):
	result={}
	varlist=[]
	setCondition=" "
	whereCondition=''
	for each in dictVO:
		if each==primary_key:
			whereCondition=' where '+each+'=%s'
		setCondition+= each+'=%s'+','
		varlist.append(dictVO[each])
	setCondition=setCondition[:len(setCondition)-1]
	result['setCondition']=setCondition
	result['whereCondition']=whereCondition
	result['vars']=varlist
	return result

def create_update_sqlVO(model,dictVO):
	primary_key=get_primary_key(model)
	condition=create_update_condition(dictVO,primary_key)
	whereCondition=condition.get('whereCondition')
	if(whereCondition==''):
		print('传入的属性没有提供主键属性，故无法更新')
		return
	setCondition=condition.get('setCondition')
	if(model!=None):
		table_name = model._meta.db_table
	sql='UPDATE '+table_name+' SET '+setCondition
	return {'sql':sql,'vars':condition.get('vars')}

#DELETE
def create_delete_condition(dictVO,primary_key):
	result={}
	varlist=[]
	whereCondition=''
	for each in dictVO:
		if each==primary_key:
			whereCondition=' where '+each+'=:'+each
			varlist.append(dictVO[each])
	result['whereCondition']=whereCondition
	result['vars']=varlist
	return result
#‘DELETE FROM TABBLE_NAME WHERE ’
def create_delete_sqlVO(model,dictVO):
	primary_key=get_primary_key(model)
	condition=create_update_condition(dictVO,primary_key)
	whereCondition=condition.get('whereCondition')
	if(whereCondition==''):
		print('传入的属性没有提供主键属性，故无法删除')
		return
	if(model!=None):
		table_name = model._meta.db_table
	sql='DELETE FROM '+table_name+ whereCondition
	return {'sql':sql,'vars':condition.get('vars')}








'''
我为oracle数据库专门写一个
'''
#INSERT
def create_ora_insert_condition(dictVO):
	result={}
	varlist=[]
	columnCondition=" ("
	valueCondition=" ("
	for each in dictVO:
		columnCondition+=each+','
		valueCondition+=":"+each+","
	columnCondition=columnCondition[:len(columnCondition)-1]+')'
	valueCondition=valueCondition[:len(valueCondition)-1]+')'
	result['columnCondition']=columnCondition
	result['valueCondition']=valueCondition
	result['vars']=dictVO
	return result

def create_ora_insert_sqlVO(model,dictVO):
	condition = create_ora_insert_condition(dictVO)
	columnCondition=condition.get('columnCondition')
	valueCondition=condition.get('valueCondition')
	if(model!=None):
		table_name = model._meta.db_table
	sql ='INSERT INTO '+table_name+columnCondition+' VALUES'+valueCondition
	return {'sql':sql,'vars':condition.get('vars')}
#SELECT
def create_ora_select_condition(dictVO):
	result={}
	varlist=[]
	whereCondition=" where "
	for each in dictVO:
		whereCondition+=each+'=:'+each+' and '
	whereCondition=whereCondition[:len(whereCondition)-5]
	print('{0}'.format(whereCondition))
	result['whereCondition']=whereCondition
	result['vars']=dictVO
	return result

def create_ora_select_sqlVO(model,dictVO):
	condition = create_ora_select_condition(dictVO)
	whereCondition=condition.get('whereCondition')
	if(model!=None):
		table_name = model._meta.db_table
	sql ='SELECT * FROM '+table_name+whereCondition
	print(sql)
	return {'sql':sql,'vars':condition.get('vars')}

#update
def create_ora_update_condition(dictVO,primary_key):
	result={}
	varlist=[]
	setCondition=" "
	whereCondition=''
	print(primary_key)
	for each in dictVO:
		print(each)
		print(each==primary_key)
		if each==primary_key:
			whereCondition=' where '+each+'=:'+each
			continue
		setCondition+= each+'=:'+each+','
		varlist.append(dictVO[each])
	setCondition=setCondition[:len(setCondition)-1]
	result['setCondition']=setCondition
	result['whereCondition']=whereCondition
	result['vars']=dictVO
	return result

def create_ora_update_sqlVO(model,dictVO):
	primary_key=get_primary_key(model)
	condition=create_ora_update_condition(dictVO,primary_key)
	whereCondition=condition.get('whereCondition')
	if(whereCondition==''):
		print('传入的属性没有提供主键属性，故无法更新')
		return
	setCondition=condition.get('setCondition')
	if(model!=None):
		table_name = model._meta.db_table
	sql='UPDATE '+table_name+' SET '+setCondition+whereCondition
	return {'sql':sql,'vars':condition.get('vars')}

#DELETE
def create_ora_delete_condition(dictVO,primary_key):
	result={}
	varlist=[]
	whereCondition=''
	whereValue={}
	for each in dictVO:
		if each==primary_key:
			whereCondition=' where '+each+'=:'+each
			whereValue[each]=dictVO[each]
			continue
		
	result['whereCondition']=whereCondition
	result['vars']=whereValue
	return result
#‘DELETE FROM TABBLE_NAME WHERE ’
def create_ora_delete_sqlVO(model,dictVO):
	primary_key=get_primary_key(model)
	condition=create_ora_delete_condition(dictVO,primary_key)
	whereCondition=condition.get('whereCondition')
	if(whereCondition==''):
		print('传入的属性没有提供主键属性，故无法删除')
		return
	if(model!=None):
		table_name = model._meta.db_table
	sql='DELETE FROM '+table_name+ whereCondition
	return {'sql':sql,'vars':condition.get('vars')}

'''
取出近一个月的关联字段
'''
def batch_import_data(path,model,procedurename):
	attrs = get_model_attrs(model)
	entities=[]
	#records=model.objects.get_all_tr()
	wb2 = openpyxl.load_workbook(path)
	sheet_names = wb2.get_sheet_names()
	for sheet_name in sheet_names:
		ws = wb2[sheet_name]
		j=1
		entity={}
		own_uid=''
		for row in ws.rows:
			if(j==1):
				own_uid=row[j].value
				j+=1
				continue
			elif(j==2):
				j+=1
				continue
			print(row)
			entity['own_uid']=own_uid
			entity['procedure']=procedurename
			for i in range(len(row)):
				#print('{0}:{1}'.format(attrs[i+1],row[i].value))
				if row[i].value is not None:
					entity[attrs[i+1]]=row[i].value.strip()
			entities.append(entity)
			entity={}
		print('***********************************{0}'.format(entities))
		for entity in entities:
			m=model()
			m.set_attr(attr = entity)
			m.save()
		entities=[]

def proper_uid_data():
	sqlVO={}
	#select_uid_sql="select HEATNO from QG_IFL_USER.IF_BOF_L2L2_POOL t where MSG_DATE between  to_date(to_char(sysdate-30,'yyyy-mm-dd'),'yyyy-mm-dd') and to_date(to_char(sysdate,'yyyy-mm-dd'),'yyyy-mm-dd')"
	select_uid_sql="select HEATNO from QG_IFL_USER.IF_BOF_L2L2_POOL t "
	db_name='l2'
	sqlVO={'sql':select_uid_sql,'db_name':db_name}
	return sqlVO



def get_value_by_uid_sqlVO(record,own_id):
	sqlVO={}
	select_uid_sql = 'SELECT ' +record['from_col'].upper() +' FROM ' +record['from_table'].upper()+' where '+record['from_uid'].upper()+'=%s'
	db_name=record['from_system'].lower()
	varlist=[own_id]
	sqlVO={'sql':select_uid_sql,'db_name':db_name,'vars':varlist}
	#print(sqlVO)
	return sqlVO

def create_update_by_uid_sqlVO(dictVO):
	sqlVO={}
	#dictVO={'record':record,'value':insert_value,'table_name':table_name,'own_id':own_id}
	record=dictVO['record']
	#print('dictVO:{0}'.format(dictVO))
	#print('record:{0}'.format(record))
	select_uid_sql = 'UPDATE '+dictVO['table_name']+' SET ' +record['own_col'] +'=%s'+' WHERE '+ record['own_uid']+'=%s'
	sqlVO={'sql':select_uid_sql,'vars':[dictVO['value'][0][record['from_col'].upper()],dictVO['own_id']]}
	return sqlVO

#多个数据组合到一起类型问题
def change_value_2_insert(value):
	return value

#从指定的表中取出关联字段的数据并去重
def drop_multi_uid(uids,reference_uid,record):
	print('***************************{0}'.format(len(uids)))
	unique_set=set()
	for uid in uids:
		print(uid)
		unique_set.add(uid[reference_uid])
	print(len(unique_set))
	unique_uids_list=list(unique_set)
	unique_uids=[]
	uid_dict={}
	for uid in unique_uids_list:
		uid_dict[record['from_uid'].upper()]=uid
		unique_uids.append(uid_dict)
		uid_dict={}
	return unique_uids
'''
复合主键相关工具方法
'''
def import_multikey_file(path,model,procedurename):
	attrs = get_model_attrs(model)
	entities=[]
	#records=model.objects.get_all_tr()
	wb2 = openpyxl.load_workbook(path)
	sheet_names = wb2.get_sheet_names()
	for sheet_name in sheet_names:
		ws = wb2[sheet_name]
		j=1
		entity={}
		own_uid1=''
		own_uid2=''
		for row in ws.rows:
			if(j==1):
				own_uid1=row[j].value.strip()
				own_uid2=row[j+1].value.strip()
				print('own_uid1:{0},own_uid2:{1}'.format(own_uid1,own_uid2))
				j+=1
				continue
			elif(j==2):
				j+=1
				continue
			entity['own_uid1']=own_uid1
			entity['own_uid2']=own_uid2
			entity['procedure']=procedurename
			print(attrs)
			for i in range(len(row)):
				#print('{0}:{1}'.format(attrs[i+1],row[i].value))
				if row[i].value is not None:
					entity[attrs[i+2]]=row[i].value.strip()
			entities.append(entity)
			entity={}
		print(entities)
		for entity in entities:
			m=model()
			m.set_attr(attr = entity)
			m.save()
		entities=[]

def get_multikey_2uid_data(record):
	sqlVO={}
	select_uid_sql = 'SELECT ' +record['from_uid1'].upper() +','+ record['from_uid2'].upper()+' FROM ' +record['from_table'].upper() +' WHERE ROWNUM<%s'
	db_name=record['from_system'].lower()
	varlist=[ROW_NUM]
	sqlVO={'sql':select_uid_sql,'db_name':db_name,'vars':varlist}
	return sqlVO

def insert_2uid_sqlVO(uid,record):
	sqlVO={}
	sql="INSERT INTO unique_col(relation_col1,relation_col2) VALUES(%s,%s)"
	varlist=[uid[record['from_uid1'].upper()],uid[record['from_uid2'].upper()]]
	sqlVO={'sql':sql,'vars':varlist}
	return sqlVO

#去除重复2uids
def unique_uids_sqlVO():
	sqlVO={}
	sql="SELECT distinct relation_col1,relation_col2 FROM unique_col"
	sqlVO={'sql':sql}
	return sqlVO

def delete_table(table_name):
	sqlVO={}
	sql='DELETE FROM '+table_name
	sqlVO={'sql':sql}
	return sqlVO

def get_single_uid_data(record):
	sqlVO={}
	select_uid_sql = 'SELECT distinct ' +record['from_uid'].upper()+' FROM ' +record['from_table'].upper() +' WHERE ROWNUM<%s'
	db_name=record['from_system'].lower()
	varlist=[ROW_NUM]
	sqlVO={'sql':select_uid_sql,'db_name':db_name,'vars':varlist}
	return sqlVO
	
def insert_uid_sqlVO(uid,record):
	sqlVO={}
	sql="INSERT INTO unique_col(relation_col1) VALUES(%s)"
	varlist=[uid[record['from_uid'].upper()]]
	sqlVO={'sql':sql,'vars':varlist}
	return sqlVO
	
def unique_single_uids_sqlVO():
	sqlVO={}
	sql="SELECT distinct relation_col1 FROM unique_col"
	sqlVO={'sql':sql}
	return sqlVO

def get_value_by_2uid_sqlVO(record,own_uid1,own_uid2):
	sqlVO={}
	select_uid_sql = 'SELECT ' +record['from_col'] +' FROM ' +record['from_table']+' where '+record['from_uid1'].upper()+'=%s'+' and '+record['from_uid2'].upper()+'=%s'
	db_name=record['from_system'].lower()
	varlist=[own_uid1,own_uid2]
	sqlVO={'sql':select_uid_sql,'db_name':db_name,'vars':varlist}
	return sqlVO

def create_update_by_2uid_sqlVO(dictVO):
	sqlVO={}
	#dictVO={'record':record,'value':insert_value,'table_name':table_name,'own_id':own_id}
	record=dictVO['record']
	print('dictVO:{0}'.format(dictVO))
	print('record:{0}'.format(record))
	select_uid_sql = 'UPDATE '+dictVO['table_name']+' SET ' +record['own_col'] +'=%s'+' WHERE '+record['own_uid1']+'=%s'+' and '+record['own_uid2']+'=%s'
	sqlVO={'sql':select_uid_sql,'vars':[dictVO['value'][0][record['from_col'].upper()],dictVO['own_uid1'],dictVO['own_uid2']]}
	return sqlVO

def create_sqlVO(sql,varlist):
	sqlVO={}
	sqlVO['sql']=sql
	sqlVO['vars']=varlist
	return sqlVO

def write_2_excel(ele_list,procedurename):
	basepath='\\data_import\\media\\upload\\'+procedurename+'数据缺失分析.xlsx'
	filepath=BASE_DIR+basepath
	wb = openpyxl.Workbook()
	ws = wb.active
	for ele in ele_list:
		eles=[ele[key] for key in ele]
		ws.append(eles)
	wb.save(filepath)
	return filepath

def create_get_records_sqlVO(procedurename,table_name):
	sqlVO={}
	select_records_sql = 'select * from '+table_name+' where `procedure`=%s'
	varlist=[procedurename]
	sqlVO={'sql':select_records_sql,'vars':varlist}
	return sqlVO

def create_delete_records_sqlVO(procedurename,table_name):
	sqlVO={}
	delete_records_sql = 'delete from '+table_name+' where `procedure`=%s'
	varlist=[procedurename]
	sqlVO={'sql':delete_records_sql,'vars':varlist}
	return sqlVO

def create_sum_sqlVO(table_name):
	sqlVO={}
	sum_sql = 'select count(id) from '+table_name
	sqlVO={'sql':sum_sql}
	return sqlVO

#根据表名查询数据库每个字段缺失情况
def create_get_columns_sqlVO(table_name,db_name):
	sqlVO={}
	columns_sql = 'select * from '+table_name
	sqlVO['sql']=columns_sql
	sqlVO['db_name']=db_name
	return sqlVO

def create_get_sum_sqlVO(col,table_name,db_name):
	sqlVO={}
	sql = 'select count('+col+') from '+table_name
	sqlVO['sql']=sql
	sqlVO['db_name']=db_name
	return sqlVO

def create_sqlVO_by_dbname(sql,varlist,db_name):
	sqlVO={}
	sqlVO['sql']=sql
	sqlVO['vars']=varlist
	sqlVO['db_name']=db_name
	return sqlVO

def write_ana_2_file(tblist,tables_info):
	#['铁前','炼钢','轧钢']
	basepath='\\data_import\\media\\upload\\QG_USER.PRO_LF_HIS_CHRGDGEN.xlsx'
	filepath=BASE_DIR+basepath
	wb = openpyxl.Workbook()
	ws = wb.active
	title=["表名","数据总量","字段名","空值数","空值率","零值数","零值率"]
	ws.append(title)
	for table_name in tblist:
		table_info=tables_info.get(table_name)
		if(table_info==None):
			ws.append([table_name])
			ws.append([])
			continue;
		total_num=table_info.get("total_num")
		if(total_num==0):
			ws.append([table_name,total_num])
			ws.append([])
			continue;
		columns=table_info.get("columns")
		for column in columns[0][1:]:
			column_info=table_info.get(column)
			null_num=column_info.get("null_num")
			null_rate=round(column_info.get("null_rate"),2)
			zero_num=column_info.get("zero_num",None)
			zero_rate=column_info.get("zero_rate",None)
			if zero_rate!=None:
				zero_rate=round(zero_rate,2)
			write_line=[table_name,total_num,column,null_num,null_rate,zero_num,zero_rate]
			ws.append(write_line)
		ws.append([])
	wb.save(filepath)
